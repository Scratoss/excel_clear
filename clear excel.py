import re
import shutil

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager

# Загружаем входной файл
input_table = '/Users/egoryurihin/Downloads/Боевая таблица за 2015 год.xlsx'

# Определяем место расположения выходного файла
output_table = '/Users/egoryurihin/Downloads/Новая Боевая таблица за 2015 год.xlsx'

# Создаем выходной файл с метаданными путем копирования входного
shutil.copyfile(input_table, output_table)

# Создаем список страниц входного файла
sheet_names = openpyxl.load_workbook(input_table).sheetnames

# Запускаем браузер без графики
options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('--disable-notifications')
options.add_argument("--disable-gpu")
driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=options)

# Задаем адрес страницы
driver.get("https://www.kody.su/check-tel#text")

# Находим форму ввода
elem = driver.find_element_by_name("/html/body/div/div[2]/div/div[1]/div/form/input[1]")

# Создаем счетчик для отсчета итераций при ожидании
time_1 = len(sheet_names)-1

# Создаем список с наименованием столбцов
c_names = ['ID', 'Дата', 'Запрос/Клиент', 'Менеджер', 'Источник информации о нас', 'Чем интересуется', 'Итог',
               'Прим.', 'Повторные контакты', 'Данные', 'опроса', 'клиента', 'Комментарии']

# Стартуем постраничную обработку файла
for s_name in sheet_names:
    # Считываем страницу
    df = pd.read_excel(input_table, sheet_name=s_name)
    columns_count = len(df.columns)
    df = df.drop([0])
    # Присваиваем столбцам названия
    c_names_1 = []
    for i in range(0, columns_count):
        c_names_1.append(c_names[i])
    df.columns = c_names_1

    # Заполняем пропуски нулями (для определения только рабочих строк таблицы)
    df['ID'] = df['ID'].fillna(0)
    df = df.loc[df['ID'] != 0]

    # Заполняем отсутствующие данные
    df = df.fillna('!Данные не указаны!')

    # Забираем данные на разбор
    client = df['Запрос/Клиент'], str.split(',')

    # Вытаскиваем почту
    e_mails = []
    for e in client[0]:
        if '@' in e:
            e_mails.append(re.findall('[a-zA-z0-9\.]+[@][a-zA-z0-9\S]+[.][a-zA-z]*\w', e))
        else:
            e_mails.append('0')
    for i in range(0, len(e_mails)):
        if e_mails[i] == []:
            e_mails[i] = '0'
        if e_mails[i] == '0':
            e_mails[i] = '!Почта не указана!'
    e_mail = []
    for i in range(0, len(e_mails)):
        if e_mails[i] != '!Почта не указана!':
            e_mail.append(e_mails[i][0])
        else:
            e_mail.append(e_mails[i])
    df['Электронная почта'] = e_mail # Добавляем новый столбец и вносим в него почтовые адреса

    #  Добавляем вспомогательный определитель региона  по интернет
    internet_region = []
    for i in client[0]:
        if '.ru' in i:
            internet_region.append('ru')
        else:
            internet_region.append('not')


    # Обрабатываем номера телефонов
    phones = []
    for i in client[0]:
        phones.append(re.findall('[0-9+]*\d',i))
    phone=[]
    for i in phones:
        phone.append(''.join(i))
    for i in range(0,len(phone)):
        if len(phone[i])<10:
            phone[i]='0'
        if '+'in phone[i]:
            phone[i]= phone[i].split('+')
    for i in phone:
        if len(i)>1 and i[0]=='':
            i.remove(i[0])
    for i in range(0,len(phone)):
        if type (phone[i])== list:
            if len(phone[i][0])>10:
                phone[i]=phone[i][0][:11]
            else:
                phone[i]=phone[i][1][:11]
        if type(phone[i])==list and len(phone[i])==1:
            phone[i]=phone[i][0]
        if phone[i][0]=='8' and len(phone[i])>10:
            phone[i]=phone[i].replace('8','7',1)
        if phone[i] =='0':
            phone[i]='!Телефон не указан!'
    for i in range(0,len(phone)):
        if phone[i][0]!='7' and phone[i][0]!='3' and internet_region[i]=='ru' and phone[i]!='!Телефон не указан!':
            phone[i]='7'+phone[i]
        elif phone[i][0]=='3' and internet_region[i]=='ru' and phone[i]!='!Телефон не указан!':
            phone[i]='7'+phone[i]

    for i in range(0,len(phone)):
        if phone[i] != '!Телефон не указан!' and internet_region[i]=='ru':
                phone[i]=re.findall('[7][0-9]{10}',phone[i])
                if phone[i]!=[]:
                    phone[i]=phone[i][0]



    # Определяем по телефону регион путем заполнения формы на сайте https://www.kody.su/check-tel#text и считывания региона
    region_name = []
    city_name = []

    for i in phone:
        if i != '!Телефон не указан!' and len(i) > 10:
            elem.send_keys(i[:11])
            elem.send_keys(Keys.ENTER)
            try:
                region = driver.find_element_by_xpath('/html/body/div/div[2]/div/div[1]/div/table/tbody/tr[2]/td[1]/strong')
                city = driver.find_element_by_xpath('/html/body/div/div[2]/div/div[1]/div/table/tbody/tr[2]/td[2]/strong')
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div/div[1]/div/form/input[1]")
                elem.clear()
                region_name.append(region)
                city_name.append(city)
            except:
                region_name.append('Ошибка: Номер не найден')
                city_name.append('Ошибка: Номер не найден')
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div/div[1]/div/form/input[1]")
                elem.clear()
                continue
        else:
            region_name.append('!Регион не определен!')
            city_name.append('!Регион не определен!')
    for i in range(0, len(region_name)):
        if region_name[i] == '':
            region_name[i] = '!Регион не определен!'

    df['Регион'] = region_name # Создаем новый столбец со страной
    df['Город'] = city_name #Создаем новый столбец с регионом

    # Сохраняем данные в файл в новую страницу
    book = load_workbook(output_table)
    writer = pd.ExcelWriter(output_table, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, s_name)
    writer.save()

    # Выводим сообщение для информирования пользователя
    print('Ожидайте...',time_1)
    time_1 = time_1 - 1

print('!!!Форматирование завершено!!!')
driver.close()
